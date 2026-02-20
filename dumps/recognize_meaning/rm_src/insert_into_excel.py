#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Скрипт для добавления данных из rm_output_data.json в Excel файл.
Для каждой школы (сверяя по id) добавляет столбцы из overall_school_metrics.
"""

import json
import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Any, List

# Пути к файлам
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))

JSON_FILE = os.path.join(SCRIPT_DIR, "..", "rm_data", "rm_output", "rm_output_data.json")
EXCEL_FILE = os.path.join(PROJECT_ROOT, "global_data", "Здания школ.xlsx")
SHEET_NAME = "schools_2_stage"


def load_json_data(json_file: str) -> List[Dict[str, Any]]:
    """Загружает данные из JSON файла и возвращает overall_school_metrics"""
    print(f"Загрузка данных из JSON: {json_file}")
    
    if not os.path.exists(json_file):
        raise FileNotFoundError(f"JSON файл не найден: {json_file}")
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    if 'overall_school_metrics' not in data:
        raise KeyError("В JSON файле отсутствует поле 'overall_school_metrics'")
    
    metrics = data['overall_school_metrics']
    print(f"Загружено метрик для школ: {len(metrics)}")
    
    return metrics


def load_excel_data(excel_file: str, sheet_name: str) -> pd.DataFrame:
    """Загружает данные из Excel файла"""
    print(f"Загрузка данных из Excel: {excel_file}")
    print(f"Лист: {sheet_name}")
    
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Excel файл не найден: {excel_file}")
    
    # Проверяем наличие листа
    wb = load_workbook(excel_file, read_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Лист '{sheet_name}' не найден в файле. Доступные листы: {wb.sheetnames}")
    wb.close()
    
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    print(f"Загружено записей из Excel: {len(df)}")
    print(f"Колонок в Excel: {len(df.columns)}")
    
    return df


def rename_columns_for_analyst(df: pd.DataFrame) -> pd.DataFrame:
    """Переименовывает столбцы в более понятные названия для аналитика"""
    rename_dict = {}
    
    for col in df.columns:
        if col == 'school_id':
            continue  # Пропускаем school_id, он будет удален или уже есть в Excel
        
        # Обрабатываем столбцы вида topic_<тема>_<метрика>
        if col.startswith('topic_'):
            parts = col.split('_')
            if len(parts) >= 3:
                topic = '_'.join(parts[1:-1])  # Все части между 'topic' и метрикой
                metric = parts[-1]  # Последняя часть - метрика
                
                # Переводим метрики в понятные названия
                metric_names = {
                    'cnt': 'Количество_отзывов',
                    'neg_share': 'Доля_негативных_проц',
                    'pos_cnt': 'Количество_позитивных',
                    'neg_cnt': 'Количество_негативных',
                    'sentiment': 'Тональность'
                }
                
                # Переводим темы в понятные названия (с заглавной буквы)
                topic_names = {
                    'администрация': 'Администрация',
                    'буллинг': 'Буллинг',
                    'еда': 'Еда',
                    'инфраструктура': 'Инфраструктура',
                    'охрана': 'Охрана',
                    'ремонт': 'Ремонт',
                    'уборка': 'Уборка',
                    'учителя': 'Учителя'
                }
                
                # Формируем новое название
                topic_display = topic_names.get(topic, topic.capitalize())
                metric_display = metric_names.get(metric, metric)
                
                new_name = f"{topic_display}_{metric_display}"
                rename_dict[col] = new_name
    
    # Применяем переименование
    if rename_dict:
        df = df.rename(columns=rename_dict)
        print(f"Переименовано столбцов: {len(rename_dict)}")
    
    return df


def merge_data(df: pd.DataFrame, metrics: List[Dict[str, Any]]) -> pd.DataFrame:
    """Объединяет данные из Excel и JSON по school_id"""
    print("\nОбъединение данных...")
    
    # Создаем DataFrame из метрик
    metrics_df = pd.DataFrame(metrics)
    
    # Переименовываем столбцы для аналитика
    metrics_df = rename_columns_for_analyst(metrics_df)
    
    # Преобразуем school_id в строку для сопоставления
    metrics_df['school_id'] = metrics_df['school_id'].astype(str)
    
    # Находим колонку с ID в Excel (может быть 'ID', 'id', 'school_id' и т.д.)
    id_columns = ['ID', 'id', 'school_id', 'ID школы', 'Школа ID']
    excel_id_col = None
    
    for col in id_columns:
        if col in df.columns:
            excel_id_col = col
            break
    
    if excel_id_col is None:
        # Если не нашли стандартные колонки, ищем колонку, содержащую 'id' в названии
        for col in df.columns:
            if 'id' in str(col).lower():
                excel_id_col = col
                break
    
    if excel_id_col is None:
        raise ValueError(f"Не найдена колонка с ID в Excel. Доступные колонки: {list(df.columns)}")
    
    print(f"Используется колонка ID в Excel: '{excel_id_col}'")
    
    # Преобразуем ID в Excel в строку для сопоставления
    df[excel_id_col] = df[excel_id_col].astype(str)
    
    # Объединяем данные по ID
    # Используем left join, чтобы сохранить все строки из Excel
    merged_df = df.merge(
        metrics_df,
        left_on=excel_id_col,
        right_on='school_id',
        how='left',
        suffixes=('', '_from_json')
    )
    
    # Удаляем дублирующуюся колонку school_id из JSON (если она совпадает с excel_id_col)
    if 'school_id' in merged_df.columns and excel_id_col != 'school_id':
        merged_df = merged_df.drop(columns=['school_id'])
    
    # Подсчитываем статистику объединения
    # Используем все столбцы из metrics_df кроме school_id
    metric_columns = [col for col in metrics_df.columns if col != 'school_id']
    if metric_columns:
        matched_count = merged_df[metric_columns].notna().any(axis=1).sum()
        print(f"Сопоставлено школ: {matched_count} из {len(df)}")
    
    return merged_df


def save_to_excel(df: pd.DataFrame, excel_file: str, sheet_name: str):
    """Сохраняет DataFrame в Excel файл"""
    print(f"\nСохранение данных в Excel...")
    print(f"Файл: {excel_file}")
    print(f"Лист: {sheet_name}")
    
    try:
        # Используем pd.ExcelWriter с режимом 'a' (append) для добавления/замены листа
        # if_sheet_exists='replace' заменяет существующий лист
        with pd.ExcelWriter(excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"✓ Данные успешно сохранены")
        print(f"✓ Всего записей: {len(df)}")
        print(f"✓ Всего колонок: {len(df.columns)}")
        
    except PermissionError:
        print(f"\n[ERROR] Ошибка доступа к файлу!")
        print(f"Файл '{excel_file}' открыт в другой программе (Excel, LibreOffice и т.д.).")
        print(f"Пожалуйста, закройте файл и попробуйте снова.")
        raise
    except Exception as e:
        print(f"\n[ERROR] Ошибка при сохранении файла: {e}")
        raise


def main():
    """Основная функция"""
    print("=" * 80)
    print("ДОБАВЛЕНИЕ ДАННЫХ ИЗ JSON В EXCEL")
    print("=" * 80)
    
    try:
        # Загружаем данные из JSON
        metrics = load_json_data(JSON_FILE)
        
        # Загружаем данные из Excel
        df = load_excel_data(EXCEL_FILE, SHEET_NAME)
        
        # Объединяем данные
        merged_df = merge_data(df, metrics)
        
        # Сохраняем результат
        save_to_excel(merged_df, EXCEL_FILE, SHEET_NAME)
        
        print("=" * 80)
        print("ОБРАБОТКА ЗАВЕРШЕНА УСПЕШНО")
        print("=" * 80)
        
    except Exception as e:
        print(f"\n[ERROR] Ошибка при выполнении: {e}")
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    main()

