import json
import os
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Any, List

# Пути к файлам
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = 'C:/repos/analysis-of-educational-institutions/global_data/Здания школ.xlsx'
SHEET_NAME = "schools_1_stage"
JSON_FILE = os.path.join(SCRIPT_DIR, "..", "vam_data", "school.json")

# Проверка существования листа
wb = load_workbook(EXCEL_FILE, read_only=True)
if SHEET_NAME not in wb.sheetnames:
    wb.close()
    raise ValueError(f"Лист '{SHEET_NAME}' не найден в файле. Доступные листы: {wb.sheetnames}")
wb.close()

# Чтение данных из Excel
df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)

# Создание структуры данных для JSON
schools_data = {
    "schools": []
}

# Обработка каждой строки
for index, row in df.iterrows():
    school_id = int(row['ID']) if pd.notna(row['ID']) else None
    school_name = str(row['Название на 2ГИС']) if pd.notna(row['Название на 2ГИС']) else ""
    school_short_name = str(row['Короткое название']) if pd.notna(row['Короткое название']) else ""
    school_adres = str(row['Адрес']) if pd.notna(row['Адрес']) else ""
    
    school_entry = {
        "school_id": school_id,
        "school_name": school_name,
        "school_short_name": school_short_name,
        "school_adres": school_adres
    }
    
    schools_data["schools"].append(school_entry)
    print(f"Обработана школа: ID={school_id}, Название={school_name}")

# Запись данных в JSON файл
with open(JSON_FILE, 'w', encoding='utf-8') as f:
    json.dump(schools_data, f, ensure_ascii=False, indent=4)

print(f"\nДанные успешно записаны в файл: {JSON_FILE}")
print(f"Всего обработано школ: {len(schools_data['schools'])}")