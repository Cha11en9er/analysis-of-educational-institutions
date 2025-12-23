import json
import os
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
from typing import Dict, Any, List
import psycopg2

# Загрузка переменных окружения
load_dotenv()

# Параметры подключения к БД
DB_CONFIG = {
    'host': os.getenv('DB_HOST'),
    'port': os.getenv('DB_PORT'),
    'database': os.getenv('DB_NAME'),
    'user': os.getenv('DB_USER'),
    'password': os.getenv('DB_PASSWORD')
}


def insert_data(school_id: int, school_name: str, school_adres: str):

    conn = psycopg2.connect(**DB_CONFIG)
    cursor = conn.cursor()

    cursor.execute("INSERT INTO ca.school (school_id, school_name, school_adres) VALUES (%s, %s, %s)", (school_id, school_name, school_adres))
    conn.commit()
    cursor.close()
    conn.close()

# Пути к файлам
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.abspath(os.path.join(SCRIPT_DIR, "..", ".."))

JSON_FILE = os.path.join(SCRIPT_DIR, "..", "rm_data", "rm_output", "rm_output_data.json")
EXCEL_FILE = 'C:/repos/analysis-of-educational-institutions/global_data/Здания школ.xlsx'
SHEET_NAME = "schools_2_stage"

wb = load_workbook(EXCEL_FILE, read_only=True)
if SHEET_NAME not in wb.sheetnames:
    wb.close()
    raise ValueError(f"Лист '{SHEET_NAME}' не найден в файле. Доступные листы: {wb.sheetnames}")
wb.close()

df = pd.read_excel(EXCEL_FILE, sheet_name=SHEET_NAME)
# у меня только id, название и адрес

for index, row in df.iterrows():
    id = row['ID']
    name = row['Название на 2ГИС']
    adres = row['Адрес']
    print(id, name, adres)

    # вставляем данные в базу данных
    insert_data(school_id=id, school_name=name, school_adres=adres)